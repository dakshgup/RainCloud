from django.http import HttpResponseRedirect
from django.shortcuts import render
from django import forms
from django import forms
from collections import OrderedDict
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.plotting import figure, output_file, show
import sqlite3
import openpyxl
import form

# Create your views here.


def load_data(file_path):
    all_data = pd.read_excel(file_path, sheet_name=None)

    for i in range(len(all_data)):
        try:
            all_data[str(i + 1)] = all_data[str(i + 1)].drop(
                columns=[
                    "Unnamed: 0",
                    "Unnamed: 3",
                    "Unnamed: 7",
                    "Unnamed: 8",
                    "Unnamed: 9",
                    "Unnamed: 10",
                    "Unnamed: 11",
                    "Unnamed: 12",
                    "Unnamed: 13",
                    "Unnamed: 14",
                    "Unnamed: 15",
                ]
            )[8:]
            all_data[str(i + 1)] = (
                all_data[str(i + 1)]
                .rename(
                    columns={
                        "Unnamed: 1": "title",
                        "Unnamed: 2": "max_value",
                        "Unnamed: 4": "time",
                        "Unnamed: 5": "load",
                        "Unnamed: 6": "temp",
                    }
                )
                .reset_index()
            )

        except KeyError:
            pass
    # print(all_data.keys())
    return all_data


june_2019 = "C:/Users/daksh/PycharmProjects/aleo_dataviz_1/Aleo I - June Daily Data Report 2020.xls"
june_2020 = "C:/Users/daksh/PycharmProjects/aleo_dataviz_1/Aleo I - June Daily Data Report 2020.xls"

# print(load_data(june_2019)['1'][['Unnamed: 1', 'Unnamed: 2']])


def get_feature(sheet, date, feature, unit):
    data = load_data(sheet)[str(date)]
    print(type(data))
    if unit == 1:
        feature_map = {
            "winding_temp": 1,
            "bearing_temp": 2,
            "generator_volt": 3,
            "generator_current": 4,
            "line_voltage": 5,
            "line_load": 6,
            "line_average_load": 7,
        }

    elif unit == 2:
        feature_map = {
            "winding_temp": 16,
            "bearing_temp": 17,
            "generator_volt": 18,
            "generator_current": 19,
            "line_voltage": 20,
            "line_load": 21,
            "line_average_load": 22,
        }
    # [['max_value', 'time', 'load', 'temp']]
    # [feature_map[feature]]
    # print(feature_map[feature])
    out = data.loc[feature_map[feature]][["max_value", "time", "load", "temp"]]
    out = pd.DataFrame(out).transpose()

    return out


def get_day(sheet, date, unit):
    data = load_data(sheet)[str(date)]
    if unit == 1:
        out = data.loc[1:7][["max_value", "time", "load", "temp"]]
        out = pd.DataFrame(out).transpose()
        out = out.rename(
            columns={
                1: "winding_temp",
                2: "bearing_temp",
                3: "generator_volt",
                4: "generator_current",
                5: "line_voltage",
                6: "line_load",
                7: "line_average_load",
            }
        )
    if unit == 2:
        out = data.loc[16:22][["max_value", "time", "load", "temp"]]
        out = pd.DataFrame(out).transpose()
        out = out.rename(
            columns={
                16: "winding_temp",
                17: "bearing_temp",
                18: "generator_volt",
                19: "generator_current",
                20: "line_voltage",
                21: "line_load",
                22: "line_average_load",
            }
        )

    return out


# print(get_day(june_2019, 8, 1))


def plot_single_month(reading, param, freq, unit):
    if freq == "daily":
        y = []
        for i in range(1, 30):
            data = get_day(june_2020, i, unit)
            # print(data.columns)
            y.append(data[reading].loc[param])
    # plt.plot(y)
    # plt.title(reading + " " + param + " " + freq + " for unit " + str(unit))
    # plt.show()
    x = list(range(1, 30))
    title = reading + " " + param + " " + freq + " unit: " + str(unit)
    output_file("lines.html", title="line plot example")
    plot = figure(title=title, x_axis_label="Day of Month", width=800, height=600)
    plot.line(x, y)
    show(plot)


# plot_single_month('bearing_temp', 'max_value', "daily", 2)
# plot_single_month('line_load', 'max_value', "daily", 2)


def get_month(sheet, unit):
    days = {}
    for i in range(31):
        try:
            days[i] = get_day(sheet, i, unit)
        except:
            days[i] = np.nan
    return days


# print(get_month(june_2020, 1))


def write_to_excel(data):
    date = str(data["date"][0])
    # 2020-07-18
    print(date)
    year = str(int(date[0:4]))
    month = str(int(date[5:7]))
    day = str(int(date[8:10]))
    plant = str(data["plant_options"][0])
    wkbTempName = "aleo_web_app/static/data/{}/{}_template.xlsx".format(plant, plant)
    fileSaveName = plant + "_" + year + "_" + month + ".xlsx"
    wbkSaveName = "aleo_web_app/static/data/{}/{}/{}".format(plant, year, fileSaveName)
    try:
        wbk = openyxl.load_workbook(wbkSaveName)
    except:
        wbk = openpyxl.load_workbook(wkbTempName)

    ws = wbk[day]

    ws.cell(7, 6).value = date

    ########## UNIT I ####################################
    ws.cell(11, 3).value = int(data["u1_wt_max_value"][0])
    ws.cell(12, 3).value = int(data["u1_bt_max_value"][0])
    ws.cell(13, 3).value = int(data["u1_gv_max_value"][0])
    ws.cell(14, 3).value = int(data["u1_gc_max_value"][0])
    ws.cell(15, 3).value = int(data["u1_lv_max_value"][0])
    ws.cell(16, 3).value = int(data["u1_lo_max_value"][0])
    ws.cell(17, 3).value = int(data["u1_avg_load"][0])

    ws.cell(11, 4).value = int(data["u1_wt_channel"][0])
    ws.cell(12, 4).value = int(data["u1_bt_channel"][0])

    ws.cell(11, 3).value = int(data["u1_wt_time"][0])
    ws.cell(12, 3).value = int(data["u1_bt_time"][0])
    ws.cell(13, 3).value = int(data["u1_gv_time"][0])
    ws.cell(14, 3).value = int(data["u1_gc_time"][0])
    ws.cell(15, 3).value = int(data["u1_lv_time"][0])
    ws.cell(16, 3).value = int(data["u1_lo_time"][0])

    ws.cell(11, 3).value = int(data["u1_wt_load"][0])
    ws.cell(12, 3).value = int(data["u1_bt_load"][0])
    ws.cell(13, 3).value = int(data["u1_gv_load"][0])
    ws.cell(14, 3).value = int(data["u1_gc_load"][0])
    ws.cell(15, 3).value = int(data["u1_lv_load"][0])
    ws.cell(16, 3).value = int(data["u1_lo_load"][0])

    ws.cell(11, 3).value = int(data["u1_wt_ambient"][0])
    ws.cell(12, 3).value = int(data["u1_bt_ambient"][0])
    ws.cell(13, 3).value = int(data["u1_gv_ambient"][0])
    ws.cell(14, 3).value = int(data["u1_gc_ambient"][0])
    ws.cell(15, 3).value = int(data["u1_lv_ambient"][0])
    ws.cell(16, 3).value = int(data["u1_lo_ambient"][0])

    ws.cell(19, 3).value = int(data["u1_gv_min"][0])
    ws.cell(20, 3).value = int(data["u1_lo_min"][0])

    ws.cell(22, 3).value = int(data["u1_at_max"])
    ws.cell(22, 4).value = int(data["u1_at_min"])

    ws.cell(22, 3).value = int(data["u1_at_min"])
    ws.cell(22, 4).value = int(data["u1_at_min"])

    ws.cell(19, 5).value = int(data["u1_gv_min_time"])
    ws.cell(19, 6).value = int(data["u1_gv_min_load"])
    ws.cell(19, 7).value = int(data["u1_gv_min_at"])

    ws.cell(20, 5).value = int(data["u1_lo_min_time"])
    ws.cell(20, 6).value = int(data["u1_lo_min_load"])
    ws.cell(20, 7).value = int(data["u1_lo_min_at"])

    ########## UNIT II ###################################
    ws.cell(26, 3).value = int(data["u2_wt_max_value"][0])
    ws.cell(27, 3).value = int(data["u2_bt_max_value"][0])
    ws.cell(28, 3).value = int(data["u2_gv_max_value"][0])
    ws.cell(29, 3).value = int(data["u2_gc_max_value"][0])
    ws.cell(30, 3).value = int(data["u2_lv_max_value"][0])
    ws.cell(31, 3).value = int(data["u2_lo_max_value"][0])
    ws.cell(32, 3).value = int(data["u2_avg_load"][0])

    ws.cell(26, 4).value = int(data["u2_wt_channel"][0])
    ws.cell(27, 4).value = int(data["u2_bt_channel"][0])

    ws.cell(26, 3).value = int(data["u2_wt_time"][0])
    ws.cell(27, 3).value = int(data["u2_bt_time"][0])
    ws.cell(28, 3).value = int(data["u2_gv_time"][0])
    ws.cell(29, 3).value = int(data["u2_gc_time"][0])
    ws.cell(30, 3).value = int(data["u2_lv_time"][0])
    ws.cell(31, 3).value = int(data["u2_lo_time"][0])

    ws.cell(26, 3).value = int(data["u2_wt_load"][0])
    ws.cell(27, 3).value = int(data["u2_bt_load"][0])
    ws.cell(28, 3).value = int(data["u2_gv_load"][0])
    ws.cell(29, 3).value = int(data["u2_gc_load"][0])
    ws.cell(30, 3).value = int(data["u2_lv_load"][0])
    ws.cell(31, 3).value = int(data["u2_lo_load"][0])

    ws.cell(26, 3).value = int(data["u2_wt_ambient"][0])
    ws.cell(27, 3).value = int(data["u2_bt_ambient"][0])
    ws.cell(28, 3).value = int(data["u2_gv_ambient"][0])
    ws.cell(29, 3).value = int(data["u2_gc_ambient"][0])
    ws.cell(30, 3).value = int(data["u2_lv_ambient"][0])
    ws.cell(31, 3).value = int(data["u2_lo_ambient"][0])

    ws.cell(34, 3).value = int(data["u2_gv_min"][0])
    ws.cell(35, 3).value = int(data["u2_lo_min"][0])

    ws.cell(37, 3).value = int(data["u2_at_max"][0])
    ws.cell(37, 4).value = int(data["u2_at_min"][0])

    ws.cell(34, 5).value = int(data["u2_gv_min_time"][0])
    ws.cell(34, 6).value = int(data["u2_gv_min_load"][0])
    ws.cell(34, 7).value = int(data["u2_gv_min_at"][0])

    ws.cell(35, 5).value = int(data["u2_lo_min_time"][0])
    ws.cell(35, 6).value = int(data["u2_lo_min_load"][0])
    ws.cell(35, 6).value = int(data["u2_lo_min_at"][0])

    ########### WEATHER ######################################
    # TODO: this stuff.

    ########### PRODUCTION ###################################
    ws.cell(39, 3).value = int(data["u1_prod"][0])
    ws.cell(39, 4).value = int(data["u2_prod"][0])

    ws.cell(39, 7).value = int(data["cumul_monthly"][0])
    ws.cell(39, 8).value = int(data["cumul_total"][0])
    wbk.save(wbkSaveName)


def homepage(request):
    return render(request, "base.html")


def form(request):
    return render(request, "form.html")


def analyze(request):
    return render(request, "analyze.html")


def submission(request):

    dicto = dict(request.GET)
    write_to_excel(dicto)

    """
    f = open("aleo_web_app/static/data/test.txt", "w")
    f.write(str(dicto["date"]) + "\n")
    f.write(str(dicto["plant_options"]) + "\n")
    f.write(str(dicto["clear"]) + "\n")
    f.write(str(dicto["snow"]))
    """
    return render(
        request,
        "base.html",
        {"message": "Successfully submitted form for " + str(dicto["date"])},
    )
